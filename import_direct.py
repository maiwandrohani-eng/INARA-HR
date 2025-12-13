#!/usr/bin/env python3
import openpyxl
import psycopg2
from datetime import datetime, date
import uuid

# Database connection
DB_CONFIG = {
    'dbname': 'inara_hris',
    'user': 'inara_user',
    'password': 'inara_password',
    'host': 'localhost',
    'port': 5432
}

def convert_date(date_str):
    """Convert various date formats to date object"""
    if not date_str or str(date_str).strip() == '':
        return None
    
    date_str = str(date_str).strip()
    
    # Try DD/MM/YYYY format
    for sep in ['/', '-', '.']:
        if sep in date_str:
            try:
                parts = date_str.split(sep)
                if len(parts) == 3:
                    day = int(parts[0])
                    month = int(parts[1])
                    year = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
                    return date(year, month, day)
            except:
                pass
    
    return None

def main():
    print("Connecting to database...")
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    print("Reading Excel file...")
    wb = openpyxl.load_workbook('/Users/maiwand/INARA-HR/employees_import.xlsx')
    ws = wb.active
    
    success_count = 0
    failed_count = 0
    errors = []
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Process each row (skip header)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Found {len(rows)} employees to import\n")
    
    for i, row in enumerate(rows, 1):
        if not row[0] or not row[1]:  # Skip if no first/last name
            failed_count += 1
            continue
        
        try:
            # Extract values
            first_name = str(row[0]).strip()
            last_name = str(row[1]).strip()
            work_email = str(row[2]).strip() if row[2] else f"{first_name.lower()}.{last_name.lower()}@inara.org"
            phone = str(row[4]).strip() if row[4] else None
            mobile = str(row[5]).strip() if row[5] else None
            date_of_birth = convert_date(row[6]) or date(1990, 1, 1)
            gender = str(row[7]).strip().lower() if row[7] else 'male'
            nationality = str(row[8]).strip() if row[8] else 'Unknown'
            
            # Employment type mapping
            emp_type = str(row[9]).strip().upper().replace(' ', '_') if row[9] else 'FULL_TIME'
            if 'CONSULTANT' in emp_type or 'CONSULTANCY' in emp_type:
                employment_type = 'CONSULTANT'
            elif 'PART' in emp_type:
                employment_type = 'PART_TIME'
            elif 'VOLUNTEER' in emp_type:
                employment_type = 'VOLUNTEER'
            else:
                employment_type = 'FULL_TIME'
            
            work_location = str(row[10]).strip() if row[10] else 'Unknown'
            hire_date = convert_date(row[11]) or date.today()
            
            # Determine country code from work location or nationality
            location = work_location.upper()
            nat = nationality.upper()
            if 'EGYPT' in location or 'CAIRO' in location or 'EGYPT' in nat:
                country_code = 'EG'
            elif 'LEBANON' in location or 'BEIRUT' in location or 'LEBANON' in nat:
                country_code = 'LB'
            elif 'PALESTINE' in location or 'GAZA' in location or 'PALESTINE' in nat:
                country_code = 'PS'
            elif 'SYRIA' in location or 'DAMASCUS' in location or 'SYRIA' in nat:
                country_code = 'SY'
            elif 'TURK' in location or 'ISTANBUL' in location or 'ANKARA' in location or 'TURK' in nat:
                country_code = 'TR'
            elif 'AFGHANISTAN' in location or 'KABUL' in location or 'AFGHANISTAN' in nat:
                country_code = 'AF'
            elif 'PAKISTAN' in location or 'PAKISTAN' in nat:
                country_code = 'PK'
            elif 'GEORGIA' in location or 'GEORGIA' in nat:
                country_code = 'GE'
            elif 'UK' in nat or 'BRITAIN' in nat or 'ENGLAND' in nat:
                country_code = 'GB'
            elif 'FRANCE' in location or 'FRANCE' in nat:
                country_code = 'FR'
            else:
                country_code = 'XX'  # Unknown
            
            employee_id = str(uuid.uuid4())
            employee_number = f"INR-{i:04d}"
            
            # Insert query with all required fields
            insert_query = """
            INSERT INTO employees (
                id, employee_number, first_name, last_name, work_email,
                phone, mobile, date_of_birth, gender, nationality,
                employment_type, work_location, hire_date, status,
                country_code, is_deleted, created_at, updated_at
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, FALSE, NOW(), NOW()
            )
            """
            
            cursor.execute(insert_query, (
                employee_id, employee_number, first_name, last_name, work_email,
                phone, mobile, date_of_birth, gender, nationality,
                employment_type, work_location, hire_date, 'ACTIVE',
                country_code
            ))
            
            conn.commit()
            success_count += 1
            print(f"[{i}/{len(rows)}] ✓ {first_name} {last_name}")
            
        except Exception as e:
            conn.rollback()
            failed_count += 1
            error_msg = str(e)[:100]
            print(f"[{i}/{len(rows)}] ✗ {first_name} {last_name} - {error_msg}")
            errors.append({'name': f"{first_name} {last_name}", 'error': error_msg})
    
    cursor.close()
    conn.close()
    
    print("\n" + "=" * 50)
    print("Import Summary:")
    print(f"✓ Successfully added: {success_count}")
    print(f"✗ Failed: {failed_count}")
    print("=" * 50)
    
    if errors and len(errors) <= 10:
        print("\nErrors:")
        for idx, err in enumerate(errors, 1):
            print(f"{idx}. {err['name']}: {err['error']}")

if __name__ == "__main__":
    main()
